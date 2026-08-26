"""Extract the current NAIC VM-20 Appendix 2 workbook set into generic structured-table JSON.

Raw workbooks remain ignored under data/work. The tracked output retains workbook,
sheet, and cell provenance and is review-only until an independent table review.
"""

from __future__ import annotations

import hashlib
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = REPO_ROOT / "data" / "work" / "structured-table-sources" / "vm20-appendix2-2026"
OUTPUT_PATH = REPO_ROOT / "data" / "processed" / "structured_tables" / "vm20-appendix2-tables.json"
OFFICIAL_PAGE = "https://content.naic.org/pbr_data.htm"
MANUAL_URL = "https://content.naic.org/sites/default/files/pbr_data_valuation_manual_current_edition.pdf"
MANUAL_SHA256 = "496cab9f387c84971df69eab1528d93aea70f7e57c8429661f2765498b38d4e9"
RETRIEVED_AT = "2026-08-26T00:00:00.000Z"

SOURCE_DEFINITIONS = {
    "table-a-workbook": {
        "filename": "pbr-2025-vm20-table-a-baseline-annual-default-costs.xlsx",
        "url": "https://content.naic.org/sites/default/files/pbr-2025-vm20-table-a-baseline-annual-default-costs.xlsx",
        "workbookScope": "Table A",
    },
    "table-f-g-workbook": {
        "filename": "pbr-2026-vm20-table-f-g-current-spreads.xlsx",
        "url": "https://content.naic.org/sites/default/files/pbr-2026-vm20-table-f-g-current-spreads.xlsx",
        "workbookScope": "Tables F and G",
    },
    "table-h-i-workbook": {
        "filename": "pbr-2026-vm20-table-h-i-long-term-spreads.xlsx",
        "url": "https://content.naic.org/sites/default/files/pbr-2026-vm20-table-h-i-long-term-spreads.xlsx",
        "workbookScope": "Tables H and I",
    },
    "table-j-workbook": {
        "filename": "pbr-2026-vm20-table-j-swaps.xlsx",
        "url": "https://content.naic.org/sites/default/files/pbr-2026-vm20-table-j-swaps.xlsx",
        "workbookScope": "Table J",
    },
    "table-k-workbook": {
        "filename": "pbr_data_table_k_conversion.xlsx",
        "url": "https://content.naic.org/sites/default/files/inline-files/pbr_data_table_k_conversion.xlsx",
        "workbookScope": "Table K",
    },
}

MANUAL_CITATIONS = {
    "A": ("20-91–20-92", "135–136"),
    "B": ("20-92", "136"),
    "C": ("20-92", "136"),
    "D": ("20-92", "136"),
    "E1": ("20-93", "137"),
    "E2": ("20-93", "137"),
    "F": ("20-93–20-94", "137–138"),
    "G": ("20-93–20-94", "137–138"),
    "H": ("20-94–20-95", "138–139"),
    "I": ("20-94–20-95", "138–139"),
    "J": ("20-95–20-96", "139–140"),
    "K": ("20-91, 20-96", "135, 140"),
}

INVENTORY_TITLES = {
    "A": "Baseline Annual Default Costs",
    "B": "Baseline Default Cost Margin",
    "C": "Empirical CTE 70 Default Rates",
    "D": "Prescribed Cumulative Default Rates",
    "E1": "Sorted Recovery Data, CTE 70 Recovery Rates, and Implied Margin",
    "E2": "Final Recovery Rates by PBR Credit Rating",
    "F": "Investment Grade Current Benchmark Spreads",
    "G": "Below Investment Grade Current Benchmark Spreads",
    "H": "Investment Grade Long-Term Benchmark Spreads",
    "I": "Below Investment Grade Long-Term Benchmark Spreads",
    "J": "Current and Long-Term Benchmark Swap Spreads",
    "K": "Conversion to PBR Numeric Rating",
}


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def relative(path: Path) -> str:
    return path.relative_to(REPO_ROOT).as_posix()


def display_value(value: Any, number_format: str) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        if "0.00" in number_format:
            return f"{value:.2f}"
        return str(value)
    return str(value)


def parse_title_date(title: str) -> str:
    match = re.search(r"\((\d{1,2})/(\d{1,2})/(\d{4})\)", title)
    if not match:
        raise ValueError(f"No snapshot date in title: {title}")
    return datetime(int(match.group(3)), int(match.group(1)), int(match.group(2))).date().isoformat()


def citation(source_artifact_id: str, sheet: str, cell_range: str, table_label: str) -> dict[str, Any]:
    source = SOURCE_DEFINITIONS[source_artifact_id]
    printed, physical = MANUAL_CITATIONS[table_label]
    return {
        "sourceArtifactId": source_artifact_id,
        "workbookUrl": source["url"],
        "sheetName": sheet,
        "cellRange": cell_range,
        "manualPrintedPageRange": printed,
        "manualPhysicalPageRange": physical,
    }


def value_record(ws, cell_address: str, column_id: str, column_label: str, unit: str | None) -> dict[str, Any]:
    cell = ws[cell_address]
    value = cell.value
    record: dict[str, Any] = {
        "columnId": column_id,
        "columnLabel": column_label,
        "valueType": "null" if value is None else "number" if isinstance(value, (int, float)) and not isinstance(value, bool) else "string",
        "displayValue": display_value(value, cell.number_format),
        "sourceCell": cell_address,
        "numberFormat": cell.number_format,
        "fidelity": "exact_source_cell_value",
    }
    if unit is not None:
        record["unit"] = unit
    if value is None:
        record["nullReason"] = "not_provided_in_source_cell"
    elif record["valueType"] == "number":
        record["numericValue"] = value
    else:
        record["textValue"] = str(value)
    return record


def source_artifacts() -> list[dict[str, Any]]:
    artifacts = []
    for source_id, definition in SOURCE_DEFINITIONS.items():
        path = SOURCE_ROOT / definition["filename"]
        if not path.exists():
            raise FileNotFoundError(path)
        artifacts.append({
            "sourceArtifactId": source_id,
            "filename": definition["filename"],
            "repositoryWorkPath": relative(path),
            "officialUrl": definition["url"],
            "officialIndexUrl": OFFICIAL_PAGE,
            "retrievedAt": RETRIEVED_AT,
            "sha256": sha256(path),
            "fileSizeBytes": path.stat().st_size,
            "workbookScope": definition["workbookScope"],
            "legalDisclaimer": {
                "sheetName": "LEGAL DISCLAIMER",
                "retention": "retained_by_original_workbook_hash_and_sheet_locator",
                "extractedText": None,
                "note": "The dedicated sheet exists but does not expose disclaimer text as ordinary worksheet cell values; no text was inferred or transcribed.",
            },
        })
    return artifacts


def base_table(table_id: str, label: str, title: str, description: str, units: str | None, source_artifact_id: str, dimensions, columns, notes=None) -> dict[str, Any]:
    printed, physical = MANUAL_CITATIONS[label]
    return {
        "tableId": table_id,
        "tableLabel": label,
        "title": title,
        "description": description,
        "units": units,
        "sourceArtifactId": source_artifact_id,
        "methodologyCitation": {
            "manualUrl": MANUAL_URL,
            "manualSha256": MANUAL_SHA256,
            "printedPageRange": printed,
            "physicalPageRange": physical,
        },
        "dimensionDefinitions": dimensions,
        "columnDefinitions": columns,
        "notes": notes or [],
        "versions": [],
    }


def version(table, version_id: str, as_of_date: str | None, effective_date: str | None, currentness: str, currentness_basis: str, sheet: str, source_range: str, title_cell: str, rows):
    table["versions"].append({
        "versionId": version_id,
        "asOfDate": as_of_date,
        "effectiveDate": effective_date,
        "currentness": currentness,
        "currentnessBasis": currentness_basis,
        "sheetName": sheet,
        "sourceRange": source_range,
        "titleCell": title_cell,
        "rowCount": len(rows),
        "rows": rows,
        "citation": citation(table["sourceArtifactId"], sheet, source_range, table["tableLabel"]),
    })


def extract_a() -> dict[str, Any]:
    source_id = "table-a-workbook"
    wb = load_workbook(SOURCE_ROOT / SOURCE_DEFINITIONS[source_id]["filename"], data_only=True)
    sheet = "2025 Table A "
    ws = wb[sheet]
    columns = [{"columnId": f"wal-{wal}", "label": f"WAL {wal}", "dimensionValue": wal, "sourceHeaderCell": f"{chr(66 + wal)}5"} for wal in range(1, 11)]
    table = base_table("vm20-table-a", "A", INVENTORY_TITLES["A"], "Baseline annual default cost factors by PBR credit rating and weighted average life.", "basis_points", source_id, [{"dimensionId": "pbr_numeric_rating", "label": "PBR Numeric Rating", "lookupRequired": True}, {"dimensionId": "moodys_rating", "label": "Moody's Rating", "lookupRequired": False}], columns, [{"noteId": "table-a-effective-date", "scope": "table_version", "summary": "The workbook states that this table is effective June 30, 2026.", "sourceCell": "A28"}])
    rows = []
    for row_num in range(6, 26):
        rating = ws[f"A{row_num}"].value
        moodys = ws[f"B{row_num}"].value
        rows.append({
            "rowId": f"vm20-table-a-rating-{rating}",
            "ordinal": row_num - 5,
            "dimensions": [
                {"dimensionId": "pbr_numeric_rating", "label": "PBR Numeric Rating", "value": rating, "sourceCell": f"A{row_num}"},
                {"dimensionId": "moodys_rating", "label": "Moody's Rating", "value": moodys, "sourceCell": f"B{row_num}"},
            ],
            "values": [value_record(ws, f"{chr(67 + index)}{row_num}", f"wal-{index + 1}", f"WAL {index + 1}", "basis_points") for index in range(10)],
        })
    version(table, "vm20-table-a-effective-2026-06-30", None, "2026-06-30", "current_as_of_retrieval", "The official current-data page linked this workbook on the retrieval date; the workbook states the effective date in A28.", sheet, "A1:L28", "A1", rows)
    return table


def extract_spread_pair(source_id: str, left_label: str, right_label: str, left_kind: str, right_kind: str) -> list[dict[str, Any]]:
    wb = load_workbook(SOURCE_ROOT / SOURCE_DEFINITIONS[source_id]["filename"], data_only=True)
    sheets = [name for name in wb.sheetnames if name != "LEGAL DISCLAIMER"]
    definitions = []
    for label, side, kind in [(left_label, "left", left_kind), (right_label, "right", right_kind)]:
        rating_start = 1 if side == "left" else 11
        header_ws = wb[sheets[0]]
        start_col = 2 if side == "left" else 14
        columns = [{"columnId": f"pbr-rating-{rating}", "label": f"PBR Rating {rating} ({header_ws.cell(4, start_col + offset).value})", "dimensionValue": rating, "sourceHeaderCell": f"{chr(start_col + offset + 64)}3:{chr(start_col + offset + 64)}4"} for offset, rating in enumerate(range(rating_start, rating_start + 10))]
        table = base_table(f"vm20-table-{label.lower()}", label, INVENTORY_TITLES[label], f"{kind} benchmark spreads by weighted average life and PBR credit rating.", "basis_points", source_id, [{"dimensionId": "weighted_average_life_years", "label": "Weighted Average Life (years)", "lookupRequired": True}], columns)
        definitions.append((table, side))
    latest_date = max(parse_title_date(str(wb[sheet]["A1"].value if sheet in wb.sheetnames else "")) for sheet in sheets)
    for table, side in definitions:
        for sheet in sheets:
            ws = wb[sheet]
            title_cell = "A1" if side == "left" else "M1"
            as_of = parse_title_date(str(ws[title_cell].value))
            start_col = 2 if side == "left" else 14
            wal_col = "A" if side == "left" else "M"
            rows = []
            for row_num in range(5, 36):
                wal = ws[f"{wal_col}{row_num}"].value
                values = []
                for offset in range(10):
                    col_num = start_col + offset
                    cell_address = f"{chr(64 + col_num)}{row_num}"
                    rating = (1 if side == "left" else 11) + offset
                    rating_label = ws.cell(4, col_num).value
                    values.append(value_record(ws, cell_address, f"pbr-rating-{rating}", f"PBR Rating {rating} ({rating_label})", "basis_points"))
                rows.append({"rowId": f"{table['tableId']}-{as_of}-wal-{wal}", "ordinal": row_num - 4, "dimensions": [{"dimensionId": "weighted_average_life_years", "label": "Weighted Average Life (years)", "value": wal, "sourceCell": f"{wal_col}{row_num}"}], "values": values})
            source_range = f"A1:K35" if side == "left" else "M1:W35"
            version(table, f"{table['tableId']}-{as_of}", as_of, None, "current_as_of_retrieval" if as_of == latest_date else "historical_snapshot", "Latest dated sheet in the official workbook on the retrieval date." if as_of == latest_date else "An earlier dated sheet retained in the same official workbook.", sheet, source_range, title_cell, rows)
    return [item[0] for item in definitions]


def extract_j() -> dict[str, Any]:
    source_id = "table-j-workbook"
    wb = load_workbook(SOURCE_ROOT / SOURCE_DEFINITIONS[source_id]["filename"], data_only=True)
    sheets = [name for name in wb.sheetnames if "2026" in name]
    columns = [
        {"columnId": "current-swap-spread", "label": "Current Swap Spread", "sourceHeaderCell": "B2:B3"},
        {"columnId": "long-term-swap-spread", "label": "Long-Term Swap Spread", "sourceHeaderCell": "C2:C3"},
    ]
    notes = [
        {"noteId": "libor-to-sofr-disclosure-reference", "scope": "workbook_context", "summary": "The January sheet directs users to the LIBOR-to-SOFR disclosure sheet for calculation information; applicability beyond that source reference requires reviewer confirmation.", "sourceSheet": "January_2026", "sourceCell": "A38", "relatedSheet": "LIBOR to SOFR Disclosure"},
        {"noteId": "short-tenor-current-sofr-source", "scope": "current-swap-spread at 3M and 6M", "summary": "The January source note identifies a contracted Term SOFR data source and reserves the source provider's market-data rights.", "sourceSheet": "January_2026", "sourceCell": "A39"},
    ]
    table = base_table("vm20-table-j", "J", INVENTORY_TITLES["J"], "Current and long-term swap benchmark spreads by weighted average life or short tenor.", "basis_points", source_id, [{"dimensionId": "weighted_average_life_or_tenor", "label": "Weighted Average Life or Tenor", "lookupRequired": True}], columns, notes)
    latest_date = max(parse_title_date(str(wb[s]["A1"].value)) for s in sheets)
    for sheet in sheets:
        ws = wb[sheet]
        as_of = parse_title_date(str(ws["A1"].value))
        rows = []
        for row_num in range(4, 37):
            tenor = ws[f"A{row_num}"].value
            rows.append({
                "rowId": f"vm20-table-j-{as_of}-tenor-{str(tenor).lower()}",
                "ordinal": row_num - 3,
                "dimensions": [{"dimensionId": "weighted_average_life_or_tenor", "label": "Weighted Average Life or Tenor", "value": tenor, "sourceCell": f"A{row_num}"}],
                "values": [value_record(ws, f"B{row_num}", "current-swap-spread", "Current Swap Spread", "basis_points"), value_record(ws, f"C{row_num}", "long-term-swap-spread", "Long-Term Swap Spread", "basis_points")],
            })
        version(table, f"vm20-table-j-{as_of}", as_of, None, "current_as_of_retrieval" if as_of == latest_date else "historical_snapshot", "Latest dated sheet in the official workbook on the retrieval date." if as_of == latest_date else "An earlier dated sheet retained in the same official workbook.", sheet, "A1:C36", "A1", rows)
    return table


def extract_k() -> dict[str, Any]:
    source_id = "table-k-workbook"
    wb = load_workbook(SOURCE_ROOT / SOURCE_DEFINITIONS[source_id]["filename"], data_only=True)
    ws = wb["Table K"]
    mappings = [
        ("moodys-rating", "Moody's Rating", 3, 14),
        ("sp-rating", "S&P Rating", 4, 15),
        ("fitch-rating", "Fitch Rating", 5, 16),
        ("dbrs-rating", "DBRS Rating", 6, 17),
        ("realpoint-rating", "RealPoint Rating", 7, 18),
        ("am-best-rating", "AM Best Rating", 8, 19),
        ("naic-designation", "NAIC Designation", 9, 20),
        ("naic-commercial-mortgage-designation", "NAIC Commercial Mortgage Designation", 10, 21),
    ]
    columns = [{"columnId": item[0], "label": item[1], "sourceHeaderCell": f"A{item[2]}/A{item[3]}"} for item in mappings]
    table = base_table("vm20-table-k", "K", INVENTORY_TITLES["K"], "Mapping from agency ratings and NAIC designations to PBR numeric ratings.", None, source_id, [{"dimensionId": "pbr_numeric_rating", "label": "PBR Numeric Rating", "lookupRequired": True}, {"dimensionId": "grade_band", "label": "Grade Band", "lookupRequired": False}], columns)
    rows = []
    for numeric_rating in range(1, 21):
        investment_grade = numeric_rating <= 10
        source_col = 1 + (numeric_rating if investment_grade else numeric_rating - 10)
        col_letter = chr(64 + source_col)
        numeric_row = 11 if investment_grade else 22
        values = []
        for column_id, label, ig_row, big_row in mappings:
            values.append(value_record(ws, f"{col_letter}{ig_row if investment_grade else big_row}", column_id, label, None))
        rows.append({
            "rowId": f"vm20-table-k-rating-{numeric_rating}",
            "ordinal": numeric_rating,
            "dimensions": [
                {"dimensionId": "pbr_numeric_rating", "label": "PBR Numeric Rating", "value": numeric_rating, "sourceCell": f"{col_letter}{numeric_row}"},
                {"dimensionId": "grade_band", "label": "Grade Band", "value": "investment_grade" if investment_grade else "below_investment_grade", "sourceCell": "B2" if investment_grade else "B13"},
            ],
            "values": values,
        })
    version(table, "vm20-table-k-current-undated", None, None, "current_as_of_retrieval", "The undated workbook was linked as current Table K on the official current-data page on the retrieval date; no workbook version date was inferred.", "Table K", "A1:K22", "A1", rows)
    return table


def retrieval_units(tables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    units = []
    for table in tables:
        for table_version in table["versions"]:
            for row in table_version["rows"]:
                dimensions = "; ".join(f"{item['label']}: {item['value']}" for item in row["dimensions"])
                units.append({
                    "retrievalUnitId": f"retrieval-{row['rowId']}",
                    "tableId": table["tableId"],
                    "versionId": table_version["versionId"],
                    "rowId": row["rowId"],
                    "tableLabel": table["tableLabel"],
                    "tableTitle": table["title"],
                    "currentness": table_version["currentness"],
                    "asOfDate": table_version["asOfDate"],
                    "dimensionText": dimensions,
                    "retrievalText": f"VM-20 Table {table['tableLabel']} {table['title']}; {dimensions}; as of {table_version['asOfDate'] or 'undated current source'}; units {table['units'] or 'mapping'}.",
                    "citation": table_version["citation"],
                    "reviewOnly": True,
                })
    return units


def build() -> dict[str, Any]:
    tables = [extract_a()]
    tables.extend(extract_spread_pair("table-f-g-workbook", "F", "G", "current market", "current market"))
    tables.extend(extract_spread_pair("table-h-i-workbook", "H", "I", "long-term", "long-term"))
    tables.extend([extract_j(), extract_k()])
    units = retrieval_units(tables)
    inventory = []
    ingested = {table["tableLabel"] for table in tables}
    for label, title in INVENTORY_TITLES.items():
        printed, physical = MANUAL_CITATIONS[label]
        inventory.append({
            "tableLabel": label,
            "title": title,
            "status": "ingested_review_only" if label in ingested else "unavailable_on_current_official_page",
            "tableId": f"vm20-table-{label.lower()}" if label in ingested else None,
            "reason": "Parsed from an official workbook linked on the NAIC current-data page." if label in ingested else "The 2026 Valuation Manual describes this table, but the NAIC current-data page did not link a current workbook on the retrieval date; no values were inferred from methodology prose or historical files.",
            "manualPrintedPageRange": printed,
            "manualPhysicalPageRange": physical,
        })
    versions = [version_item for table in tables for version_item in table["versions"]]
    rows = [row for version_item in versions for row in version_item["rows"]]
    values = [value for row in rows for value in row["values"]]
    return {
        "schemaVersion": "1.0",
        "datasetId": "vm20-appendix2-structured-tables-2026-08-26",
        "title": "VM-20 Appendix 2 Structured Regulatory Tables Proof of Concept",
        "generatedAt": RETRIEVED_AT,
        "generatedBy": "scripts/extract-vm20-appendix2-tables.py",
        "governance": {
            "reviewOnly": True,
            "promotionStatus": "not_promoted",
            "learnerFacingAllowed": False,
            "appReadyAllowed": False,
            "ragReadyAllowed": False,
            "copilotExportEligible": False,
            "separateFromProseCorpus": True,
            "note": "Structured table data requires independent human review and a separate promotion decision. Prose promotion does not promote this dataset.",
        },
        "manualAuthority": {
            "sourceEditionId": "NAIC-VALUATION-MANUAL-2026",
            "sourceVersionIdentifier": "2026 NAIC Valuation Manual",
            "officialUrl": MANUAL_URL,
            "sha256": MANUAL_SHA256,
            "appendix": "VM-20 Appendix 2",
            "printedPageRange": "20-91–20-96",
            "physicalPageRange": "135–140",
            "authorityBoundary": "The manual supplies methodology and directs users to NAIC-published Tables A–K; workbook cells, not prose, are the value authority for ingested tables.",
        },
        "officialIndex": {"url": OFFICIAL_PAGE, "retrievedAt": RETRIEVED_AT, "scope": "Life PBR Data: VM-20 and VM-21"},
        "sourceArtifacts": source_artifacts(),
        "tableInventory": inventory,
        "summary": {
            "tableInventoryCount": len(inventory),
            "ingestedLogicalTableCount": len(tables),
            "unavailableLogicalTableCount": len(inventory) - len(tables),
            "tableVersionCount": len(versions),
            "currentVersionCount": sum(item["currentness"] == "current_as_of_retrieval" for item in versions),
            "historicalSnapshotCount": sum(item["currentness"] == "historical_snapshot" for item in versions),
            "rowCount": len(rows),
            "valueCount": len(values),
            "retrievalUnitCount": len(units),
        },
        "tables": tables,
        "retrievalUnits": units,
    }


def main() -> None:
    output = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else OUTPUT_PATH
    dataset = build()
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(dataset, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Built {dataset['summary']['ingestedLogicalTableCount']} logical tables, {dataset['summary']['tableVersionCount']} versions, {dataset['summary']['rowCount']} rows, and {dataset['summary']['valueCount']} values.")
    print(relative(output))


if __name__ == "__main__":
    main()
