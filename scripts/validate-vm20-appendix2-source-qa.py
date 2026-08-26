"""Verify every structured VM-20 table value against its source workbook cell."""

from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DATASET_PATH = REPO_ROOT / "data" / "processed" / "structured_tables" / "vm20-appendix2-tables.json"
OUTPUT_PATH = REPO_ROOT / "data" / "processed" / "review_packages" / "vm20-appendix2-structured-table-source-qa.json"
OUTPUT_MD_PATH = REPO_ROOT / "data" / "processed" / "review_packages" / "vm20-appendix2-structured-table-source-qa.md"


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def display_value(value: Any, number_format: str) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        if "0.00" in number_format:
            return f"{value:.2f}"
        return str(value)
    return str(value)


def main() -> None:
    dataset_path = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DATASET_PATH
    dataset = json.loads(dataset_path.read_text(encoding="utf-8"))
    artifacts = {item["sourceArtifactId"]: item for item in dataset["sourceArtifacts"]}
    workbooks = {}
    formula_workbooks = {}
    source_checks = []
    for artifact in artifacts.values():
        path = REPO_ROOT / artifact["repositoryWorkPath"]
        actual_hash = sha256(path)
        if actual_hash != artifact["sha256"]:
            raise ValueError(f"Workbook hash mismatch: {artifact['sourceArtifactId']}")
        workbooks[artifact["sourceArtifactId"]] = load_workbook(path, data_only=True)
        formula_workbooks[artifact["sourceArtifactId"]] = load_workbook(path, data_only=False)
        if artifact["legalDisclaimer"]["sheetName"] not in workbooks[artifact["sourceArtifactId"]].sheetnames:
            raise ValueError(f"Missing legal disclaimer sheet: {artifact['sourceArtifactId']}")
        source_checks.append({
            "sourceArtifactId": artifact["sourceArtifactId"],
            "filename": artifact["filename"],
            "sha256": actual_hash,
            "hashMatched": True,
            "legalDisclaimerSheetPresent": True,
        })

    checked_values = 0
    formula_cells = []
    selected_spot_checks = []
    spot_targets = {
        ("vm20-table-a", "vm20-table-a-effective-2026-06-30", "vm20-table-a-rating-1", "wal-1"),
        ("vm20-table-f", "vm20-table-f-2026-07-31", "vm20-table-f-2026-07-31-wal-1", "pbr-rating-1"),
        ("vm20-table-g", "vm20-table-g-2026-07-31", "vm20-table-g-2026-07-31-wal-30", "pbr-rating-20"),
        ("vm20-table-h", "vm20-table-h-2026-06-30", "vm20-table-h-2026-06-30-wal-1", "pbr-rating-1"),
        ("vm20-table-i", "vm20-table-i-2026-06-30", "vm20-table-i-2026-06-30-wal-30", "pbr-rating-20"),
        ("vm20-table-j", "vm20-table-j-2026-07-31", "vm20-table-j-2026-07-31-tenor-3m", "current-swap-spread"),
        ("vm20-table-k", "vm20-table-k-current-undated", "vm20-table-k-rating-9", "moodys-rating"),
    }
    for table in dataset["tables"]:
        source_id = table["sourceArtifactId"]
        workbook = workbooks[source_id]
        formula_workbook = formula_workbooks[source_id]
        for table_version in table["versions"]:
            sheet = workbook[table_version["sheetName"]]
            formula_sheet = formula_workbook[table_version["sheetName"]]
            for row in table_version["rows"]:
                for value in row["values"]:
                    source_cell = sheet[value["sourceCell"]]
                    formula_cell = formula_sheet[value["sourceCell"]]
                    if isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
                        formula_cells.append({"tableId": table["tableId"], "versionId": table_version["versionId"], "cell": value["sourceCell"], "formula": formula_cell.value})
                    if value["valueType"] == "number":
                        matched = source_cell.value == value.get("numericValue")
                    elif value["valueType"] == "string":
                        matched = str(source_cell.value) == value.get("textValue")
                    else:
                        matched = source_cell.value is None and value.get("nullReason") == "not_provided_in_source_cell"
                    display_matched = display_value(source_cell.value, source_cell.number_format) == value["displayValue"]
                    number_format_matched = source_cell.number_format == value["numberFormat"]
                    if not matched or not display_matched or not number_format_matched:
                        raise ValueError(f"Cell fidelity mismatch: {table['tableId']} {table_version['versionId']} {value['sourceCell']}")
                    checked_values += 1
                    key = (table["tableId"], table_version["versionId"], row["rowId"], value["columnId"])
                    if key in spot_targets:
                        selected_spot_checks.append({
                            "tableId": table["tableId"],
                            "versionId": table_version["versionId"],
                            "rowId": row["rowId"],
                            "columnId": value["columnId"],
                            "sheetName": table_version["sheetName"],
                            "sourceCell": value["sourceCell"],
                            "displayValue": value["displayValue"],
                            "matched": True,
                        })
    if checked_values != dataset["summary"]["valueCount"]:
        raise ValueError(f"Checked value count mismatch: {checked_values}")
    if formula_cells:
        raise ValueError(f"Unexpected formulas in structured value cells: {formula_cells[:3]}")
    if len(selected_spot_checks) != len(spot_targets):
        raise ValueError("Representative table spot-check set is incomplete.")

    result = {
        "schemaVersion": "1.0",
        "qaId": "vm20-appendix2-structured-table-source-qa-2026-08-26",
        "status": "passed",
        "datasetPath": dataset_path.relative_to(REPO_ROOT).as_posix(),
        "sourceArtifactCount": len(source_checks),
        "sourceChecks": source_checks,
        "valueCellsChecked": checked_values,
        "formulaValueCellCount": 0,
        "displayValuesChecked": checked_values,
        "numberFormatsChecked": checked_values,
        "representativeSpotChecks": selected_spot_checks,
        "legalDisclaimerSheetsRetained": len(source_checks),
        "note": "Every structured value was compared with its workbook cell using the recorded source artifact, sheet, cell address, raw value, display value, and number format. Disclaimer text was not inferred from non-cell workbook content.",
    }
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")
    lines = [
        "# VM-20 Appendix 2 Structured Table Source QA",
        "",
        "- Status: passed",
        f"- Source workbooks verified by SHA-256: {len(source_checks)}",
        f"- Structured value cells checked: {checked_values}",
        f"- Formula-backed value cells: {len(formula_cells)}",
        f"- Legal-disclaimer sheets retained by workbook hash and locator: {len(source_checks)}",
        "",
        "## Representative checks",
        "",
        "| Table | Version | Row | Column | Sheet/cell | Display value |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for check in selected_spot_checks:
        lines.append(f"| {check['tableId']} | {check['versionId']} | {check['rowId']} | {check['columnId']} | {check['sheetName']}!{check['sourceCell']} | {check['displayValue']} |")
    lines.extend(["", "This verifies machine fidelity to the retrieved workbooks; it does not replace independent human review of table identity, applicability, currentness, or legal-disclaimer presentation.", ""])
    OUTPUT_MD_PATH.write_text("\n".join(lines), encoding="utf-8")
    print(f"Verified {checked_values} structured values against {len(source_checks)} source workbooks.")


if __name__ == "__main__":
    main()
